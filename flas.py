from openpyxl import load_workbook
from flask import Flask, request, jsonify

app = Flask(__name__)

@app.route('/clean-mf', methods=['POST'])
def clean_mf():
    file = request.files['file']
    wb = load_workbook(file)
    ws = wb.active

    # Read header
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = {header: [] for header in headers if header}

    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, value in enumerate(row):
            if headers[i]:
                data[headers[i]].append(value)

    return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
