from openpyxl import load_workbook
from flask import Flask, request, jsonify

from flask import Flask, request, jsonify
import tempfile
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/clean-mf', methods=['POST'])
def clean_mf():
    uploaded_file = request.files['file']
    print("Received file:", uploaded_file.filename)

    # Save the file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        uploaded_file.save(tmp.name)
        tmp.flush()
        tmp_path = tmp.name

    wb = load_workbook(tmp_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) or []]
    print("Headers:", headers)

    data = {header: [] for header in headers if header}
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, value in enumerate(row):
            if headers[i]:
                data[headers[i]].append(value)

    print("Data extracted:", data)
    return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')